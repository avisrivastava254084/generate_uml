from collections import defaultdict

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


class WriteInExcel:

    def __init__(self, file_name='UML_Spreadsheet.xlsx'):
        self.file_name = file_name
        self.writer = pd.ExcelWriter(self.file_name, engine='xlsxwriter')

    def get_number_of_rows_in_df(self, agg_data):
        number_of_rows = 0
        for class_dict in agg_data:
            number_of_rows += 1 + len(class_dict['Methods'])
        return number_of_rows

    def create_pandas_dataframe(self, agg_data, skip_cols):
        number_of_rows = self.get_number_of_rows_in_df(agg_data)
        print('Number of rows in total: {}'.format(number_of_rows))
        columns = ['' for _ in range(skip_cols+2)]
        df = pd.DataFrame(
            index=np.arange(0, number_of_rows),
            columns=columns+['Class']
        )
        row_counter = 0
        column_counter = skip_cols-1
        # mapping to store row, col for dependents.
        dependents_col_counter = {}
        # parent: child mapping created just for plotting tree like line
        parent_to_child_mapping = defaultdict(list)
        dependee_to_dependents_mapping = defaultdict(list)
        class_row_mapping = {}  # class: row number mapping
        self.dark_edges_column = defaultdict(list)
        self.inheritance_edges_column = defaultdict(list)
        prev_class_row_counter = 0
        for class_data in agg_data:
            base_class = class_data['Class']
            class_row_mapping[base_class] = row_counter
            methods = class_data['Methods']
            parents = class_data['Parents']
            dependents = class_data['Dependents']
            columns = ['' for _ in range(skip_cols+2)]
            df.loc[row_counter] = columns + [base_class]
            row_counter += 1
            for method in methods:
                print('Inserting method: {} of class: {} at row: {} and column: {}'.format(
                    method, base_class, row_counter, skip_cols+1))
                df.iloc[row_counter, skip_cols+2] = method
                row_counter += 1
            if parents:
                for parent in parents:
                    parent_to_child_mapping[parent].append(base_class)
            if dependents:
                for dependent in dependents:
                    dependee_to_dependents_mapping[base_class].append(
                        dependent)
            prev_class_row_counter = row_counter

        for dependent in dependents_col_counter.keys():
            for column in dependents_col_counter[dependent]:
                df.iloc[class_row_mapping[dependent], column] = "←"
                self.dark_edges_column[column].append(
                    class_row_mapping[dependent])
        # get a single list of dependees and parents so as to draw a common vertical pipe
        dependees_and_parents_combined = set()
        for dependee in dependee_to_dependents_mapping.keys():
            dependees_and_parents_combined.add(dependee)
        for parent in parent_to_child_mapping.keys():
            dependees_and_parents_combined.add(parent)
        for class_ in dependees_and_parents_combined:
            df.iloc[class_row_mapping[class_]][column_counter] = "→"
            self.dark_edges_column[column_counter].append(
                class_row_mapping[class_])
            if class_ in dependee_to_dependents_mapping:
                df.iloc[class_row_mapping[class_]][skip_cols] = "→"
                for dependent in dependee_to_dependents_mapping[class_]:
                    df.iloc[class_row_mapping[dependent], column_counter] = "←"
                    self.dark_edges_column[column_counter].append(
                        class_row_mapping[dependent])
            if class_ in parent_to_child_mapping:
                df.iloc[class_row_mapping[class_]][skip_cols+1] = "▷"
                for child in parent_to_child_mapping[class_]:
                    df.iloc[class_row_mapping[child], column_counter] = "◁"
                    self.dark_edges_column[column_counter].append(
                        class_row_mapping[child])
            column_counter -= 1

        # convert all NaN to None.
        df = df.replace(np.nan, '', regex=True)
        print("Create DF was called.")
        print(df)
        return df

    def write_df_to_excel(self, df, sheet_name, skip_cols):
        df.to_excel(self.writer, sheet_name=sheet_name,
                    header=True, index=False)
        self.writer.save()
        self.writer.close()
        wb = load_workbook(filename=self.file_name)
        ws = wb[sheet_name]
        ws.column_dimensions['{}'.format(get_column_letter(
            skip_cols+1))].width = 8.4  # len(Parent) = (6+1)*1.2
        # len(Methods/Children) + 1(M) + 1(C) + 1(/) = 19*1.2
        ws.column_dimensions['{}'.format(
            get_column_letter(skip_cols+2))].width = 22.8
        bd = Side(style='thick', color='000000')
        bd_inheritance = Side(style='thick', color='FF0000')
        # to check whether a col in sheet's columns has arrived for dark edges.
        col_check_counter = 0
        for _ in ws.columns:
            if col_check_counter in self.dark_edges_column:
                column_letter = get_column_letter(col_check_counter+1)
                row_range = sorted(self.dark_edges_column[col_check_counter])
                # plus one because dataframe's 0 is excel'1 but column headings are on 1 so we need to start from 2.
                # the second addition of changing 0-indexed to 1-indexed comes in the next line while plotting.
                # Now, second index of list is +2 so that we can cover the full range(0 to 3 is to be covered,
                # so for loop would be from range(0,4))
                # get the first and last row to draw the dark edges.
                for row_iterator in range(row_range[0]+1, row_range[-1]+2):
                    print('{}{}'.format(column_letter, row_iterator+1))
                    ws['{}{}'.format(column_letter, row_iterator+1)
                       ].border = Border(left=bd)
            col_check_counter += 1
        wb.save(self.file_name)
        print("{} done!".format(sheet_name))
